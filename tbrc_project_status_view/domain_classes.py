from sqlalchemy import *
from sqlalchemy.ext.declarative import declarative_base, DeclarativeMeta
from sqlalchemy.orm.attributes import get_history
from sqlalchemy.orm import relationship, interfaces, backref
from sqlalchemy.sql import func, and_, case
from sqlalchemy.engine.row import Row
from sqlalchemy import create_engine, event
from flask_login import UserMixin

Base = declarative_base()

from wtforms import Form, StringField, IntegerField, BooleanField, DateTimeField, FieldList, FormField
from wtforms_sqlalchemy.fields import QuerySelectField, QuerySelectMultipleField

from openpyxl import Workbook
from flask import make_response, flash

from tbrc_project_status_view import session, labels, current_user, env, send_one_email

from tbrc_project_status_view import get_system_var


import hashlib, string, random, datetime, uuid, os, math, operator, logging, json, re, calendar
from .akuro_runtime import PasswordField, CommaSeparatedField, FreezeMe, TextToJSON, RandomString, TextToCOLS, HourColumn, HourField, NullableBooleanField
from . import akuro_runtime





def akuro_fix_unicode_for_json(data):
    return data.replace("\r","").replace("\n","").replace("\t","").replace("\\","\\\\").replace("\"","\\\"").replace("\r\n","").replace("\0","")

def akuro_to_json(obj, origin = "__json__"):
    if isinstance(obj.__class__, DeclarativeMeta):
        fields_json = []
        dict_origin = False
        if type(origin) is list:
            origin_2 = origin
        elif type(origin) is dict:
            origin_2 = origin.get(obj.__class__.__name__,[])
            dict_origin = True
        else:
            origin_2 = getattr(obj,origin)()
        for field in [x for x in dir(obj) if not x.startswith('_') and x != 'metadata' and x in origin_2]:
            field_value = obj.__getattribute__(field)
            if field_value is None:
                fields_json.append("\"%s\":null" % field)
            elif isinstance(field_value, bool):
                fields_json.append("\"%s\":%s" % (field,"true" if field_value else "false"))
            elif isinstance(field_value, int):
                fields_json.append("\"%s\":%s" % (field,str(field_value)))
            elif isinstance(field_value, float):
                fields_json.append("\"%s\":%s" % (field,str(field_value)))
            elif isinstance(field_value, datetime.datetime):
                fields_json.append("\"%s\":\"%s\"" % (field,str(field_value)))
            elif isinstance(field_value, datetime.time):
                fields_json.append("\"%s\":\"%s\"" % (field,str(field_value)))
            elif isinstance(field_value, str):
                fields_json.append("\"%s\":\"%s\"" % (field,akuro_fix_unicode_for_json(field_value)))            
            elif isinstance(field_value, list) or isinstance(obj.__class__, DeclarativeMeta):
                if dict_origin:
                    fields_json.append("\"%s\":%s" % (field,akuro_to_json(field_value, origin)))
                else:
                    fields_json.append("\"%s\":%s" % (field,akuro_to_json(field_value)))
            else:
                fields_json.append("\"%s\":null" % field)
        return "{"+",".join(fields_json)+"}"
    if isinstance(obj, bool):
        if obj:
            return 'true'
        return 'false'
    if isinstance(obj, int) or isinstance(obj, float):
        return str(obj)
        
    if isinstance(obj, datetime.datetime):
        return "\""+str(obj)+"\""
    
    if isinstance(obj, str):
        return "\""+akuro_fix_unicode_for_json(obj)+"\""
        
    if isinstance(obj, (list,tuple,Row)):
        fragments = [] 
        for elm in obj:
            fragments.append(akuro_to_json(elm,origin))
        return "["+",".join(fragments)+"]"
    if isinstance(obj, dict):
        fragments = [] 
        for key in obj:
            fragments.append("\""+key+"\":"+akuro_to_json(obj[key],origin))
        return "{"+",".join(fragments)+"}"
    return "null"




def process_row(index, row_data, fields, sheet):
    row_content = [None] * len(fields)
    for column in range(len(fields)):
        field_value = None
        try:
            #field_value = row_data.__getattribute__(fields[column])
            field_value = operator.attrgetter(fields[column])(row_data)
        except AttributeError as e:
            field_value = ""
        except Exception as e2:
            try:
                field_value = row_data[column]
            except Exception as e:
                field_value =  str(e) + str(e2)
        
        if field_value is None:
            pass
        elif isinstance(field_value, bool):
            true_label = "Si"
            false_label = "No"
            try:
                true_label = labels["domain_classes"][row_data.__class__.__tablename__][fields[column]]["true_label"]
                false_label = labels["domain_classes"][row_data.__class__.__tablename__][fields[column]]["false_label"]
            except:
                pass
            row_content[column] = true_label if field_value else false_label
        elif isinstance(field_value, int):
            row_content[column] = field_value
        elif isinstance(field_value, float):
            row_content[column] = field_value
        elif isinstance(field_value, str):
            row_content[column] = field_value
        elif isinstance(field_value, datetime.datetime):
            row_content[column] = field_value
        elif isinstance(field_value, list) or isinstance(field_value.__class__, DeclarativeMeta):
            row_content[column] = str(field_value)
        else:
            pass
    sheet.append(row_content)

def process_row_list(result, wb, origin, headers):
    
    if isinstance(result, list):
        if headers is None:
            headings = []
            ws1 = wb.active
            ws1.title = result[0].__class__.__name__
            try:
                org = result[0].__xlsx__()
                if origin is not None:
                    org = origin
                for col in range(len(org)):
                    title = org[col]
                    try:
                        title = labels["domain_classes"][result[0].__class__.__tablename__][org[col]]["label"]
                    except:
                        pass
                    #ws1.cell(column=col+1, row=1, value=title)
                    headings.append(title)
                ws1.append(headings)
            except:
                pass
        else:
            ws1 = wb.active
            ws1.append(headers) 
        try:
            for row in range(len(result)):
                org = result[0].__xlsx__()
                if origin is not None:
                    org = origin
                process_row(row+2,result[row],org, ws1)
        except:
            for row in range(len(result)):
                process_row(row+2,result[row],range(len(result[row])), ws1)


def akuro_to_xlsx(result, origin = None, headers = None, file_name = None):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    if len(result)>0:
        process_row_list(result, wb, origin, headers)
    tempname = ""
    if file_name is None:
        tempname = str(uuid.uuid4())+".xlsx"
    else:
        tempname = file_name+".xlsx"
    wb.save(os.path.join("temp",tempname))
    binario = open(os.path.join("temp",tempname),"rb").read()
    response = make_response(binario)
    response.headers['Content-Type'] = 'application/octet-stream'
    response.headers['Content-Disposition'] = \
                                            'inline; filename=%s' % tempname
    del result
    del wb
    return response






class Permission(Base):
    __tablename__ = 'permission'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    url = Column(String(255))
    permission_description = Column(String(255))
    
    #relations to other classes
    
    role_set = relationship(
        "Role", 
        secondary= "role_permission_association_table" , 
        back_populates="permissions_of_role"
        )
        
    def __str__(self):
        return (u''+labels["domain_classes"]["permission_system_format"]).format(self=self)
    def __json__(self):
        fields = ["url","permission_description"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["url","permission_description"]
        
        return fields
    def __freeze__(self):
        fields = ["url","permission_description"]
        fields.append("id")
        return fields
    


class Permission_Form(Form):
    url = StringField('Url')
    permission_description = StringField('Permission Description')
    
    
    
    

class Permission_List_Form(Form):
    permission_set = FieldList(FormField( Permission_Form), min_entries=1)

role_permission_association_table = Table('role_permission_association_table', Base.metadata,
    Column("role_id", Integer, ForeignKey("role.id")),
    Column("permission_id", Integer, ForeignKey("permission.id"))
)





class Role(Base):
    __tablename__ = 'role'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    name = Column(String(255))
    description = Column(String(255))
    
    #relations to other classes
    
    permissions_of_role = relationship(
        "Permission", 
        secondary= "role_permission_association_table" , 
        back_populates="role_set"
        )
        
    user_set = relationship(
        "User", 
        secondary= "user_role_association_table" , 
        back_populates="roles_of_user"
        )
        
    def __str__(self):
        return (u''+labels["domain_classes"]["role_system_format"]).format(self=self)
    def __json__(self):
        fields = ["name","description"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["name","description"]
        
        return fields
    def __freeze__(self):
        fields = ["name","description"]
        fields.append("id")
        return fields
    


class Role_Form(Form):
    name = StringField('Name')
    description = StringField('Description')
    
    def permissions_of_role_helper():
        return session.query(Permission)

    permissions_of_role = QuerySelectMultipleField("Permissions Of Role",query_factory=permissions_of_role_helper)
    
    
    

class Role_List_Form(Form):
    role_set = FieldList(FormField( Role_Form), min_entries=1)

user_role_association_table = Table('user_role_association_table', Base.metadata,
    Column("user_id", Integer, ForeignKey("user.id")),
    Column("role_id", Integer, ForeignKey("role.id"))
)
user_project_association_table = Table('user_project_association_table', Base.metadata,
    Column("user_id", Integer, ForeignKey("user.id")),
    Column("project_id", Integer, ForeignKey("project.id"))
)


class User(Base,UserMixin):
    __tablename__ = 'user'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    username = Column(String(255))
    password = Column(String(255))
    profile_picture = Column(String(255))
    user_full_name = Column(String(255))
    
    #relations to other classes
    
    roles_of_user = relationship(
        "Role", 
        secondary= "user_role_association_table" , 
        back_populates="user_set"
        )
        
    recover_token_set = relationship("RecoverToken", back_populates="token_user")
        
    projects_of_user = relationship(
        "Project", 
        secondary= "user_project_association_table" , 
        back_populates="user_set"
        )
        
    
    def get_id(self):
        return self.__getattribute__("id")
    
    def __str__(self):
        return (u''+labels["domain_classes"]["user_system_format"]).format(self=self)
        
    def hasPermision(self,url):
        for r in self.roles_of_user:
            for p in r.permissions_of_role:
                if p.url == url:
                    return True
        return False
    
    def __json__(self):
        fields = ["username","password","profile_picture","user_full_name"]
        fields.append("id")
        return fields
        
    def __xlsx__(self):
        fields = ["username","password","profile_picture","user_full_name"]
        fields.append("id")
        return fields
    
class User_Form(Form):
    username = StringField('Username')
    password = PasswordField('Password')
    profile_picture = StringField('Profile Picture')
    user_full_name = StringField('User Full Name')
    
    def roles_of_user_helper():
        return session.query(Role)

    roles_of_user = QuerySelectMultipleField("Roles Of User",query_factory=roles_of_user_helper)
    def projects_of_user_helper():
        return session.query(Project)

    projects_of_user = QuerySelectMultipleField("Projects Of User",query_factory=projects_of_user_helper)
    
    
    

class User_List_Form(Form):
    user_set = FieldList(FormField( User_Form), min_entries=1)





class RecoverToken(Base):
    __tablename__ = 'recover_token'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    token_string = Column(String(255))
    token_expires = Column(DateTime)
    @property
    def token_email(self):
        return self.token_user.username
    
    #relations to other classes
    
    token_user_id = Column(Integer, ForeignKey('user.id'))
    token_user = relationship('User', foreign_keys=[token_user_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["recover_token_system_format"]).format(self=self)
    def __json__(self):
        fields = ["token_string","token_expires","token_email"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["token_string","token_expires","token_email"]
        
        return fields
    def __freeze__(self):
        fields = ["token_string","token_expires","token_email"]
        fields.append("id")
        return fields
    


class RecoverToken_Form(Form):
    token_string = StringField('Token String')
    token_expires = DateTimeField('Token Expires', format = "%Y-%m-%d %H:%M:%S")
    
    
    def token_user_helper():
        
        return session.query(User)
        

    token_user = QuerySelectField("Token User",query_factory=token_user_helper)
    
    

class RecoverToken_List_Form(Form):
    recover_token_set = FieldList(FormField( RecoverToken_Form), min_entries=1)



class Collection(Base):
    __tablename__ = 'collection'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    collection_name = Column(String(255))
    collection_data = Column(Text())
    
    #relations to other classes
    
    def __str__(self):
        return (u''+labels["domain_classes"]["collection_system_format"]).format(self=self)
    def __json__(self):
        fields = ["collection_name","collection_data"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["collection_name","collection_data"]
        
        return fields
    def __freeze__(self):
        fields = ["collection_name","collection_data"]
        fields.append("id")
        return fields
    


class Collection_Form(Form):
    collection_name = StringField('Collection Name')
    collection_data = StringField('Collection Data')
    
    
    

class Collection_List_Form(Form):
    collection_set = FieldList(FormField( Collection_Form), min_entries=1)













class Project(Base):
    __tablename__ = 'project'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    project_name = Column(String(255))
    project_url = Column(String(255))
    project_description = Column(Text())
    project_address = Column(Text())
    project_contact = Column(String(255))
    project_invoices = Column(String(255))
    project_contract = Column(String(255))
    project_comms = Column(String(255))
    
    #relations to other classes
    
    user_set = relationship(
        "User", 
        secondary= "user_project_association_table" , 
        back_populates="projects_of_user"
        )
        
    orfi_set = relationship("Orfi", back_populates="orfi_project")
        
    budget_set = relationship("Budget", back_populates="budget_project")
        
    contact_set = relationship("Contact", back_populates="contact_project")
        
    gallery_folder_set = relationship("GalleryFolder", back_populates="gallery_folder_project")
        
    def __str__(self):
        return (u''+labels["domain_classes"]["project_system_format"]).format(self=self)
    def __json__(self):
        fields = ["project_name","project_url","project_description","project_address","project_contact","project_invoices","project_contract","project_comms"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["project_name","project_url","project_description","project_address","project_contact","project_invoices","project_contract","project_comms"]
        
        return fields
    def __freeze__(self):
        fields = ["project_name","project_url","project_description","project_address","project_contact","project_invoices","project_contract","project_comms"]
        fields.append("id")
        return fields
    


class Project_Form(Form):
    project_name = StringField('Project Name')
    project_url = StringField('Project Url')
    project_description = StringField('Project Description')
    project_address = StringField('Project Address')
    project_contact = StringField('Project Contact')
    project_invoices = StringField('Project Invoices')
    project_contract = StringField('Project Contract')
    project_comms = StringField('Project Comms')
    
    
    
    
    
    
    
    

class Project_List_Form(Form):
    project_set = FieldList(FormField( Project_Form), min_entries=1)







class Orfi(Base):
    __tablename__ = 'orfi'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    orfi_creation = Column(DateTime)
    orfi_due = Column(DateTime)
    orfi_resolved = Column(Boolean(255), default=False)
    orfi_asigned_to = Column(String(255))
    orfi_title = Column(String(255))
    
    #relations to other classes
    
    orfi_project_id = Column(Integer, ForeignKey('project.id'))
    orfi_project = relationship('Project', foreign_keys=[orfi_project_id])
        
    orfi_file_set = relationship("OrfiFile", back_populates="orfi_file_orfi")
        
    def __str__(self):
        return (u''+labels["domain_classes"]["orfi_system_format"]).format(self=self)
    def __json__(self):
        fields = ["orfi_creation","orfi_due","orfi_resolved","orfi_asigned_to","orfi_title"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["orfi_creation","orfi_due","orfi_resolved","orfi_asigned_to","orfi_title"]
        
        return fields
    def __freeze__(self):
        fields = ["orfi_creation","orfi_due","orfi_resolved","orfi_asigned_to","orfi_title"]
        fields.append("id")
        return fields
    


class Orfi_Form(Form):
    orfi_creation = DateTimeField('Orfi Creation', format = "%Y-%m-%d %H:%M:%S")
    orfi_due = DateTimeField('Orfi Due', format = "%Y-%m-%d")
    orfi_resolved = BooleanField('Orfi Resolved')
    orfi_asigned_to = StringField('Orfi Asigned To')
    orfi_title = StringField('Orfi Title')
    
    def orfi_project_helper():
        
        return session.query(Project)
        

    orfi_project = QuerySelectField("Orfi Project",query_factory=orfi_project_helper)
    
    
    

class Orfi_List_Form(Form):
    orfi_set = FieldList(FormField( Orfi_Form), min_entries=1)





class OrfiFile(Base):
    __tablename__ = 'orfi_file'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    orfi_file_name = Column(String(255))
    orfi_file_file = Column(String(255))
    
    #relations to other classes
    
    orfi_file_orfi_id = Column(Integer, ForeignKey('orfi.id'))
    orfi_file_orfi = relationship('Orfi', foreign_keys=[orfi_file_orfi_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["orfi_file_system_format"]).format(self=self)
    def __json__(self):
        fields = ["orfi_file_name","orfi_file_file"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["orfi_file_name","orfi_file_file"]
        
        return fields
    def __freeze__(self):
        fields = ["orfi_file_name","orfi_file_file"]
        fields.append("id")
        return fields
    


class OrfiFile_Form(Form):
    orfi_file_name = StringField('Orfi File Name')
    orfi_file_file = StringField('Orfi File File')
    
    def orfi_file_orfi_helper():
        
        return session.query(Orfi)
        

    orfi_file_orfi = QuerySelectField("Orfi File Orfi",query_factory=orfi_file_orfi_helper)
    
    

class OrfiFile_List_Form(Form):
    orfi_file_set = FieldList(FormField( OrfiFile_Form), min_entries=1)









class Budget(Base):
    __tablename__ = 'budget'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    budget_is_hard_cost = Column(Boolean(255), default=False)
    budget_name = Column(String(255))
    budget_company = Column(String(255))
    budget_email = Column(String(255))
    budget_phone = Column(String(255))
    budget_name2 = Column(String(255))
    budget_amount = Column(Integer(), default=0)
    budget_total_duration = Column(Integer(), default=0)
    budget_progress = Column(Integer(), default=0)
    
    #relations to other classes
    
    budget_project_id = Column(Integer, ForeignKey('project.id'))
    budget_project = relationship('Project', foreign_keys=[budget_project_id])
        
    bill_set = relationship("Bill", back_populates="bill_budget")
        
    budget_file_set = relationship("BudgetFile", back_populates="budget_file_budget")
        
    def __str__(self):
        return (u''+labels["domain_classes"]["budget_system_format"]).format(self=self)
    def __json__(self):
        fields = ["budget_is_hard_cost","budget_name","budget_company","budget_email","budget_phone","budget_name2","budget_amount","budget_total_duration","budget_progress"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["budget_is_hard_cost","budget_name","budget_company","budget_email","budget_phone","budget_name2","budget_amount","budget_total_duration","budget_progress"]
        
        return fields
    def __freeze__(self):
        fields = ["budget_is_hard_cost","budget_name","budget_company","budget_email","budget_phone","budget_name2","budget_amount","budget_total_duration","budget_progress"]
        fields.append("id")
        return fields
    


class Budget_Form(Form):
    budget_is_hard_cost = BooleanField('Budget Is Hard Cost')
    budget_name = StringField('Budget Name')
    budget_company = StringField('Budget Company')
    budget_email = StringField('Budget Email')
    budget_phone = StringField('Budget Phone')
    budget_name2 = StringField('Budget Name2')
    budget_amount = IntegerField('Budget Amount', default = 0, filters=[lambda x: x if x not in [""," ","None", None] else 0])
    budget_total_duration = IntegerField('Budget Total Duration', default = 0, filters=[lambda x: x if x not in [""," ","None", None] else 0])
    budget_progress = IntegerField('Budget Progress', default = 0, filters=[lambda x: x if x not in [""," ","None", None] else 0])
    
    def budget_project_helper():
        
        return session.query(Project)
        

    budget_project = QuerySelectField("Budget Project",query_factory=budget_project_helper)
    
    
    
    

class Budget_List_Form(Form):
    budget_set = FieldList(FormField( Budget_Form), min_entries=1)





class Bill(Base):
    __tablename__ = 'bill'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    bill_date = Column(DateTime)
    bill_invoice = Column(String(255))
    bill_amount = Column(Numeric(asdecimal=False, scale = 8, precision=20), default=0.0)
    bill_paid_to_date = Column(Numeric(asdecimal=False, scale = 8, precision=20), default=0.0)
    
    #relations to other classes
    
    bill_budget_id = Column(Integer, ForeignKey('budget.id'))
    bill_budget = relationship('Budget', foreign_keys=[bill_budget_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["bill_system_format"]).format(self=self)
    def __json__(self):
        fields = ["bill_date","bill_invoice","bill_amount","bill_paid_to_date"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["bill_date","bill_invoice","bill_amount","bill_paid_to_date"]
        
        return fields
    def __freeze__(self):
        fields = ["bill_date","bill_invoice","bill_amount","bill_paid_to_date"]
        fields.append("id")
        return fields
    


class Bill_Form(Form):
    bill_date = DateTimeField('Bill Date', format = "%Y-%m-%d")
    bill_invoice = StringField('Bill Invoice')
    bill_amount = StringField('Bill Amount', default = '0', filters=[lambda x: re.sub(r'[^0-9.]', '', str(x)) if x not in [""," ","None", None] else '0', lambda x: x if x not in [""," ","None", None] else '0' , lambda x: re.sub(r'\..*', '', x)])
    bill_paid_to_date = StringField('Bill Paid To Date', default = '0', filters=[lambda x: re.sub(r'[^0-9.]', '', str(x)) if x not in [""," ","None", None] else '0', lambda x: x if x not in [""," ","None", None] else '0' , lambda x: re.sub(r'\..*', '', x)])
    
    def bill_budget_helper():
        
        return session.query(Budget)
        

    bill_budget = QuerySelectField("Bill Budget",query_factory=bill_budget_helper)
    
    

class Bill_List_Form(Form):
    bill_set = FieldList(FormField( Bill_Form), min_entries=1)





class BudgetFile(Base):
    __tablename__ = 'budget_file'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    budget_file_name = Column(String(255))
    budget_file_file = Column(String(255))
    
    #relations to other classes
    
    budget_file_budget_id = Column(Integer, ForeignKey('budget.id'))
    budget_file_budget = relationship('Budget', foreign_keys=[budget_file_budget_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["budget_file_system_format"]).format(self=self)
    def __json__(self):
        fields = ["budget_file_name","budget_file_file"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["budget_file_name","budget_file_file"]
        
        return fields
    def __freeze__(self):
        fields = ["budget_file_name","budget_file_file"]
        fields.append("id")
        return fields
    


class BudgetFile_Form(Form):
    budget_file_name = StringField('Budget File Name')
    budget_file_file = StringField('Budget File File')
    
    def budget_file_budget_helper():
        
        return session.query(Budget)
        

    budget_file_budget = QuerySelectField("Budget File Budget",query_factory=budget_file_budget_helper)
    
    

class BudgetFile_List_Form(Form):
    budget_file_set = FieldList(FormField( BudgetFile_Form), min_entries=1)





class Contact(Base):
    __tablename__ = 'contact'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    contact_role = Column(String(255))
    contact_name = Column(String(255))
    contact_email = Column(String(255))
    contact_phone = Column(String(255))
    
    #relations to other classes
    
    contact_project_id = Column(Integer, ForeignKey('project.id'))
    contact_project = relationship('Project', foreign_keys=[contact_project_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["contact_system_format"]).format(self=self)
    def __json__(self):
        fields = ["contact_role","contact_name","contact_email","contact_phone"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["contact_role","contact_name","contact_email","contact_phone"]
        
        return fields
    def __freeze__(self):
        fields = ["contact_role","contact_name","contact_email","contact_phone"]
        fields.append("id")
        return fields
    


class Contact_Form(Form):
    contact_role = StringField('Contact Role')
    contact_name = StringField('Contact Name')
    contact_email = StringField('Contact Email')
    contact_phone = StringField('Contact Phone')
    
    def contact_project_helper():
        
        return session.query(Project)
        

    contact_project = QuerySelectField("Contact Project",query_factory=contact_project_helper)
    
    

class Contact_List_Form(Form):
    contact_set = FieldList(FormField( Contact_Form), min_entries=1)







class GalleryFolder(Base):
    __tablename__ = 'gallery_folder'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    gallery_folder_name = Column(String(255))
    gallery_folder_description = Column(String(255))
    gallery_folder_link = Column(String(255))
    
    #relations to other classes
    
    gallery_folder_project_id = Column(Integer, ForeignKey('project.id'))
    gallery_folder_project = relationship('Project', foreign_keys=[gallery_folder_project_id])
        
    gallery_item_set = relationship("GalleryItem", back_populates="gallery_item_folder")
        
    def __str__(self):
        return (u''+labels["domain_classes"]["gallery_folder_system_format"]).format(self=self)
    def __json__(self):
        fields = ["gallery_folder_name","gallery_folder_description","gallery_folder_link"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["gallery_folder_name","gallery_folder_description","gallery_folder_link"]
        
        return fields
    def __freeze__(self):
        fields = ["gallery_folder_name","gallery_folder_description","gallery_folder_link"]
        fields.append("id")
        return fields
    


class GalleryFolder_Form(Form):
    gallery_folder_name = StringField('Gallery Folder Name')
    gallery_folder_description = StringField('Gallery Folder Description')
    gallery_folder_link = StringField('Gallery Folder Link')
    
    def gallery_folder_project_helper():
        
        return session.query(Project)
        

    gallery_folder_project = QuerySelectField("Gallery Folder Project",query_factory=gallery_folder_project_helper)
    
    
    

class GalleryFolder_List_Form(Form):
    gallery_folder_set = FieldList(FormField( GalleryFolder_Form), min_entries=1)





class GalleryItem(Base):
    __tablename__ = 'gallery_item'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    gallery_item_name = Column(String(255))
    gallery_item_description = Column(String(255))
    gallery_item_link = Column(String(255))
    
    #relations to other classes
    
    gallery_item_folder_id = Column(Integer, ForeignKey('gallery_folder.id'))
    gallery_item_folder = relationship('GalleryFolder', foreign_keys=[gallery_item_folder_id])
        
    def __str__(self):
        return (u''+labels["domain_classes"]["gallery_item_system_format"]).format(self=self)
    def __json__(self):
        fields = ["gallery_item_name","gallery_item_description","gallery_item_link"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["gallery_item_name","gallery_item_description","gallery_item_link"]
        
        return fields
    def __freeze__(self):
        fields = ["gallery_item_name","gallery_item_description","gallery_item_link"]
        fields.append("id")
        return fields
    


class GalleryItem_Form(Form):
    gallery_item_name = StringField('Gallery Item Name')
    gallery_item_description = StringField('Gallery Item Description')
    gallery_item_link = StringField('Gallery Item Link')
    
    def gallery_item_folder_helper():
        
        return session.query(GalleryFolder)
        

    gallery_item_folder = QuerySelectField("Gallery Item Folder",query_factory=gallery_item_folder_helper)
    
    

class GalleryItem_List_Form(Form):
    gallery_item_set = FieldList(FormField( GalleryItem_Form), min_entries=1)



class SystemConfig(Base):
    __tablename__ = 'system_config'
    id = Column(Integer, primary_key=True)
    #basic types of fields
    freddy_emails = Column(String(255))
    system_home = Column(String(255))
    system_domain = Column(String(255))
    system_seo_default_title = Column(String(255))
    system_seo_default_image = Column(String(255))
    system_seo_default_description = Column(String(255))
    system_seo_default_keywords = Column(String(255))
    
    #relations to other classes
    
    def __str__(self):
        return (u''+labels["domain_classes"]["system_config_system_format"]).format(self=self)
    def __json__(self):
        fields = ["freddy_emails","system_home","system_domain","system_seo_default_title","system_seo_default_image","system_seo_default_description","system_seo_default_keywords"]
        fields.append("id")
        return fields
    def __xlsx__(self):
        
        fields = ["freddy_emails","system_home","system_domain","system_seo_default_title","system_seo_default_image","system_seo_default_description","system_seo_default_keywords"]
        
        return fields
    def __freeze__(self):
        fields = ["freddy_emails","system_home","system_domain","system_seo_default_title","system_seo_default_image","system_seo_default_description","system_seo_default_keywords"]
        fields.append("id")
        return fields
    


class SystemConfig_Form(Form):
    freddy_emails = StringField('Freddy Emails')
    system_home = StringField('System Home')
    system_domain = StringField('System Domain')
    system_seo_default_title = StringField('System Seo Default Title')
    system_seo_default_image = StringField('System Seo Default Image')
    system_seo_default_description = StringField('System Seo Default Description')
    system_seo_default_keywords = StringField('System Seo Default Keywords')
    
    
    

class SystemConfig_List_Form(Form):
    system_config_set = FieldList(FormField( SystemConfig_Form), min_entries=1)




class AkuroStaticCache(Base):
    __tablename__  = 'akuro_static_cache'
    id  = Column(Integer, primary_key=True)
    name =  Column(String(255))
    data =  Column(Text())

def initdb(engine):

    Base.metadata.bind = engine
    Base.metadata.create_all(engine)
    for t in Base.metadata.sorted_tables:
        result = session.execute("PRAGMA table_info(%s);"%t.name)
        columnames = []
        for row in result:
            columnames.append( row[1] )
        diference = [x for x in t.columns if x.name not in columnames]
        for diff in diference:
            is_key = False
            key_elm = None
            for elm in t.foreign_keys:
                if elm.parent == diff:
                    is_key = True
                    key_elm = elm
            if not is_key:
                session.execute("ALTER TABLE %s ADD COLUMN %s %s;"%(t.name,diff.name,diff.type.compile()))
                print ("Columna",diff.name,"Agregada")
                print ("ALTER TABLE %s ADD COLUMN %s %s;"%(t.name,diff.name,diff.type.compile()))
            else:
                session.execute("ALTER TABLE %s ADD COLUMN %s %s REFERENCES %s(id);"%(t.name,diff.name,diff.type.compile(), key_elm._column_tokens[1]))
                print ("Columna",diff.name,"Agregada")
                print ("ALTER TABLE %s ADD COLUMN %s %s REFERENCES %s(id);"%(t.name,diff.name,diff.type.compile(), key_elm._column_tokens[1]))
    session.commit()






def accessLog_before_insert(session, flush_context, instances):
    for target in session.new:
        
        
        pass
    for target in session.dirty:
        
        
        pass
    
    for target in session.deleted:
        
        
        pass

event.listen(session, 'before_flush', accessLog_before_insert)


#Desarrollo realizado por akuro SAS. Contactenos en http://akuro.co.
#Development made by akuro. contact Us at http://akuro.co
#  __    __    _______   ______    __    __     ______          _______    ______     _______                               
# /" |  | "\  /"     "| /" _  "\  /" |  | "\   /    " \        |   __ "\  /    " \   /"      \                              
#(:  (__)  :)(: ______)(: ( \___)(:  (__)  :) // ____  \       (. |__) :)// ____  \ |:        |                             
# \/      \/  \/    |   \/ \      \/      \/ /  /    ) :)      |:  ____//  /    ) :)|_____/   )                             
# //  __  \\  // ___)_  //  \ _   //  __  \\(: (____/ //       (|  /   (: (____/ //  //      /                              
#(:  (  )  :)(:      "|(:   _) \ (:  (  )  :)\        /       /|__/ \   \        /  |:  __   \                              
# \__|  |__/  \_______) \_______) \__|  |__/  \"_____/       (_______)   \"_____/   |__|  \___)                             
#                                                                                                                           
#      __       __   ___  ____  ____   _______     ______                                                                   
#     /""\     |/"| /  ")("  _||_ " | /"      \   /    " \                                                                  
#    /    \    (: |/   / |   (  ) : ||:        | // ____  \                                                                 
#   /' /\  \   |    __/  (:  |  | . )|_____/   )/  /    ) :)                                                                
#  //  __'  \  (// _  \   \\ \__/ //  //      /(: (____/ //                                                                 
# /   /  \\  \ |: | \  \  /\\ __ //\ |:  __   \ \        /                                                                  
#(___/    \___)(__|  \__)(__________)|__|  \___) \"_____/                                                                   
#                                                                                                                           
#  ______    ______    _____  ___  ___________   __       ______  ___________  _______  _____  ___      ______    ________  
# /" _  "\  /    " \  (\"   \|"  \("     _   ") /""\     /" _  "\("     _   ")/"     "|(\"   \|"  \    /    " \  /"       ) 
#(: ( \___)// ____  \ |.\\   \    |)__/  \\__/ /    \   (: ( \___))__/  \\__/(: ______)|.\\   \    |  // ____  \(:   \___/  
# \/ \    /  /    ) :)|: \.   \\  |   \\_ /   /' /\  \   \/ \        \\_ /    \/    |  |: \.   \\  | /  /    ) :)\___  \    
# //  \ _(: (____/ // |.  \    \. |   |.  |  //  __'  \  //  \ _     |.  |    // ___)_ |.  \    \. |(: (____/ //  __/  \\   
#(:   _) \\        /  |    \    \ |   \:  | /   /  \\  \(:   _) \    \:  |   (:      "||    \    \ | \        /  /" \   :)  
# \_______)\"_____/    \___|\____\)    \__|(___/    \___)\_______)    \__|    \_______) \___|\____\)  \"_____/  (_______/   
#                                                                                                                           
#  ______    ______    _____  ___  ___________   __       ______  ___________      ____  ____   ________                    
# /" _  "\  /    " \  (\"   \|"  \("     _   ") /""\     /" _  "\("     _   ")    ("  _||_ " | /"       )                   
#(: ( \___)// ____  \ |.\\   \    |)__/  \\__/ /    \   (: ( \___))__/  \\__/     |   (  ) : |(:   \___/                    
# \/ \    /  /    ) :)|: \.   \\  |   \\_ /   /' /\  \   \/ \        \\_ /        (:  |  | . ) \___  \                      
# //  \ _(: (____/ // |.  \    \. |   |.  |  //  __'  \  //  \ _     |.  |         \\ \__/ //   __/  \\                     
#(:   _) \\        /  |    \    \ |   \:  | /   /  \\  \(:   _) \    \:  |         /\\ __ //\  /" \   :)                    
# \_______)\"_____/    \___|\____\)    \__|(___/    \___)\_______)    \__|        (__________)(_______/                     
#                                                                                                                           
#      __       __   ___  ____  ____   _______     ______          ______    ______                                         
#     /""\     |/"| /  ")("  _||_ " | /"      \   /    " \        /" _  "\  /    " \                                        
#    /    \    (: |/   / |   (  ) : ||:        | // ____  \      (: ( \___)// ____  \                                       
#   /' /\  \   |    __/  (:  |  | . )|_____/   )/  /    ) :)      \/ \    /  /    ) :)                                      
#  //  __'  \  (// _  \   \\ \__/ //  //      /(: (____/ //_____  //  \ _(: (____/ //                                       
# /   /  \\  \ |: | \  \  /\\ __ //\ |:  __   \ \        /))_  ")(:   _) \\        /                                        
#(___/    \___)(__|  \__)(__________)|__|  \___) \"_____/(_____(  \_______)\"_____/                                         
#        